import requests
import Consultas 
import json
from statistics import mean
from statistics import mode
from statistics import median
import openpyxl
import matplotlib.pyplot as plt
import numpy as np


marcas = ["almay", "alva","anna sui","annabelle","benefit","boosh","burt's bees","butter london","c'est moi","cargo cosmetics","china glaze"
,"clinique","coastal classic creation","colourpop","covergirl","dalish","deciem","dior","dr. hauschka","e.l.f.","essie","fenty","glossier"
,"green people","iman","l'oreal","lotus cosmetics usa","maia's mineral galaxy","marcelle","marienatie","maybelline","milani","mineral fusion"
,"misa","mistura","moov","nudus","nyx","orly","pacifica","penny lane organics","physicians formula","piggy paint","pure anada","rejuva minerals",
"revlon","sally b's skin yummies","salon perfect","sante","sinful colours","smashbox","stila","suncoat","w3llpeople","wet n wild",
"zorah","zorah biocosmetiques"]

etiquetas =["Canadian","CertClean","Chemical Free","Dairy Free","EWG Verified","EcoCert","Fair Trade","Gluten Free","Hypoallergenic","Natural","No Talc","Non-GMO"
,"Organic","Peanut Free Product","Sugar Free","USDA Organic","Vegan","alcohol free","cruelty free","oil free","purpicks","silicone free","water free"]

tipo = ["blush","bronzer","eyebrow", "eyeliner","eyeshadow","foundation","lipstick","mascara"]  #Lista para almacenar las categorias de los maquillajes

marca_ingresada = [] #Lista para asegurar que no se repitan las marcas

marca_consulta = []


def menu():   #-----Menu con el que se elige que datos se van a consultar de la API
    print("""
A continuacion se muestran opciones para la obtencion de informacion de productos de maquillaje segun la opcion que se elija
          
          1. Consulta de la variacion de precios de distintas marcas para un maquillaje en especifico.
          
          2. Variacion de precios de los maquillajes segun su etiqueta, ej(Veganos, libres de quimicos etc.)

          """)
    while True:
        try: 
            op = int(input("Opcion: "))
            if op not in (1,2):
                print("Opcion no disponible, ingresar nuevamente")
                continue
            else:
                break
        except ValueError:
            print("La opcion ingresada tiene que ser un numero")
    return op

def valida_marca():#----------------------------------------------------Funcion que permite validar si la marca ingresada exsite en los datos
    
    print("A continuacion se presentan las marcas disponibles para la consulta de su informacion")
    input("--------------------------------------------------------------------------------------\n")
    for i in range(len(marcas)):
        print("Marcas: {}".format(marcas[i]))
    datos_consulta = input("Marca: ") 

    if datos_consulta not in marcas: 
        while True:                                             # Bucle que valida que la marca ingresada este disponible
            datos_consulta = input("Ingresa nuevamente la marca, tomando en cuenta las anteriores: ")
            if datos_consulta not in marcas:
                continue
            elif datos_consulta in marca_ingresada:
                print("Ingresa una marca diferente")
                continue
            else:
                marca_ingresada.append(datos_consulta)
                break
        return datos_consulta
    elif datos_consulta in marca_ingresada: # Valida la marca ingresada
        print("Ingresa una marca distinta")
        return valida_marca() #Vueleve a llamarse a si misma para verificar que se ingrese una marca diferente
    else:
        marca_ingresada.append(datos_consulta)
        return datos_consulta

def valida_tipo():       #------------------------------------------------------------- Funcion para validar que el tipo de maquillaje este disponible
    print("Tipos de maquillajes disponibles")
    input("---------------------------------\n")
    for i in range(len(tipo)):
        print("Tipo: {}".format(tipo[i]))
    tipo_maquillaje = input("Tipo de maquillaje: ")

    if tipo_maquillaje not in tipo:
        while True:
            tipo_maquillaje = input("Ingresa nuevamente el tipo de maquillaje, tomando en cuenta las opciones anteriores: ")
            if tipo_maquillaje not in tipo:
                continue
            else:
                break
        return tipo_maquillaje
    else:
        return tipo_maquillaje    

def valida_etiqueta():
    print("Etiquetas disponibles para su consulta ")
    input("--------------------------------------------------------------------------------------\n")
    for i in range(len(etiquetas)):
        print("Etiqueta: {}".format(etiquetas[i]))
    datos = input("Etiqueta: ")

    if datos not in etiquetas:
        while True:
            datos = input("Ingresa nuevamente la etiqueta, tomando en cuenta las anteriores: ")
            if datos not in etiquetas:
                continue
            else:
                break
    else:
        return datos
    return datos  


def cln_api(consulta):  # Funcion para filtrar la informacion de la API y faciltar la busqueda, recibe como argumento el json de la consulta de la API  #-----Miguel 
    for i in range(len(consulta)):
        consulta[i].pop("id")
        consulta[i].pop("price_sign")
        consulta[i].pop("currency")
        consulta[i].pop("image_link")
        consulta[i].pop("product_link")
        consulta[i].pop("website_link")
        consulta[i].pop("tag_list")
        consulta[i].pop("created_at")
        consulta[i].pop("updated_at")
        consulta[i].pop("product_api_url")
        consulta[i].pop("api_featured_image")
        consulta[i].pop("product_colors")
        consulta[i].pop("description")

    return consulta

def no_datos(ctrl_val):                         # Funcion para cuando se da el caso de que en alguna consulta no hay datos disponibles
    print(""" Ingresa un numero para continuar
          1. Hacer otra consulta.
          2. Salir del programa.
          """)
    while True:
        try:
            n = int(input("----> "))
            if n not in (1,2):
                print("Opcion no disponible, ingresala nuevamente ")
                continue
            else: 
                break
        except ValueError:
            print("El dato ingresado tiene que ser un numero")
    if n == 1:
        return Consultas.opciones(ctrl_val)    
    else: 
        return None
        

def save_o_del():   #Funcion para determinar si se alamcena o elimina la informacion
    print("""
          1. Almacenar informacion
          2. Descartar informacion
""")
    while True:
        try:
            x = int(input("----------------> "))
            if x not in (1,2):
                print("Elegir una opcion corrrecta ")
                continue
            else:
                break
        except ValueError:
            print("El valor ingresado tiene que ser un numero")
    return x

def n_consulta(tipo_maquillaje,lista):   #Se toman dos parametros en la funcion, el tipo y la opcion, que ya estan predefinidos
    while True:
        print("Ingresar una marca diferente para la consulta de datos: ")
        nombre_consulta = valida_marca()
        response = requests.get("http://makeup-api.herokuapp.com/api/v1/products.json?brand={}&product_type={}".format(nombre_consulta,tipo_maquillaje))
        datos = cln_api(response.json())
        if len(datos) == 0:
            marca_ingresada.remove(nombre_consulta) #Elimina de la lista el elemento del cual no se tienen datos
            print("No hay informacion disponible para el maquillaje {} de la marca {}".format(tipo_maquillaje,nombre_consulta))
            continue                                    #Regresa al principio del ciclo while
        else:
            ##---------La marca ya ha sido elegida--------------------
            precios_individuales = [] # Se guarda en la lista para calcular su promedio

            for i in range(len(datos)):
                precios_individuales.append(float(datos[i].get("price")))
            prom = mean(precios_individuales) # Se obtiene el promedio de los precios

            print("Precios:  ")
            
            print(precios_individuales)

            with open("Precios","a") as txt:
                txt.write("Marca: "+nombre_consulta+"\n")
                for i in range(len(precios_individuales)):
                    txt.write("Precio: "+str(precios_individuales[i])+"\n")
                txt.write("|-------------------------|"+"\n")
                txt.write("Promedio: "+str(prom)+"\n")
                txt.write("|-------------------------|"+"\n")

            with open("Precios_bruto","a") as txt:
                txt.write("Marca: "+nombre_consulta+"\n")
                for i in range(len(precios_individuales)):
                    txt.write(str(precios_individuales[i])+"\n")
            
            print("Promedio de precios: ",mean(precios_individuales))
            print("Moda de precios: ", mode(precios_individuales))
            print("Mediana de los precios", median(precios_individuales))

            lista.append(prom)                             #Se agrega a la lista de promedios que se regresará
            ctrl_3 = verificacion()
            if ctrl_3 == 1:
                continue
            else:
                break
    return lista


def verificacion ():  #Funcion para decidir si agregar otras marcas, en caso de que no, finaliza el programa 
    print("""
          1.Agregar marca
          2.Salir 
          """)
    while True:
        try:
            ch = int(input("------>"))
            if ch not in(1,2):
                print("Opcion no valida")
                continue
            else:
                break
        except ValueError:
            print("El dato ingresado tiene que ser un numero")
    return ch

def leer_archivo():      #Funcion para leer el archivo
    print("""
          1. Consultar informacion almacenada en el archivo
          2. Continuar
""")
    while True:
        try:
            entrada = int(input("---->"))
            if entrada not in (1,2):
                print("Opcion no disponible")
                continue
            else:
                break
        except ValueError:
            print("Dato ingresado incorrecto")
    return entrada
            
def grafica(lista_datos,caso):
    if caso == 1:
        x = np.array(marca_ingresada)
        y = np.array(lista_datos)
        plt.bar(x,y)
        
        plt.title("Variacion de precios para distintas marcas")
        plt.xlabel("Marcas")
        plt.ylabel("Precios promedio")
        plt.show() #Grafica de barras
       
        y = np.array(lista_datos)
        plt.pie(y, labels = marca_ingresada, shadow=True)
        plt.legend(title = "Marcas: ")
        plt.show() #Grafica pastel
        
        x = np.array(lista_datos)
        plt.hist(x)  #Hisotgrama
        plt.title("Histograma de la variacion de precios")
        plt.show()
    
    elif caso == 2:

        x = np.array(lista_datos)
        plt.hist(x)
        plt.title("Histograma de precios")
        plt.show()

def registro_excel(lista,caso_opc):
    if caso_opc == 1:
        libro = openpyxl.Workbook()
        pagina = libro.active
        pagina['A1'] = 'Promedio'
        for i,promedio in enumerate(lista,start=2):
            pagina.cell(row = i, column = 1, value = promedio)
        cadena = input("Nombre de la pagina: ")
        pagina.title = cadena 

        libro.save("Consulta.xlsx")

   
    elif caso_opc == 2:
        libro = openpyxl.Workbook()
        pagina = libro.active
        pagina['A1'] = "Precios"
        for i,promedio in enumerate(lista,start=2):
            pagina.cell(row = i, column = 1, value = promedio)
        cadena = input("Nombre de la pagina: ")
        pagina.title = cadena
        libro.save("Consulta.xlsx")

    else:
        pass

def leer_excel(nombre_archivo = "Consulta.xlsx"):

    print("Se leera el archivo excel: ")

    libro = openpyxl.load_workbook(nombre_archivo)

    # Selecciona la primera hoja del libro de trabajo
    hoja = libro.active

    # Obtiene los valores de la columna 'A' (promedios)
    lectura = [hoja.cell(row=i, column=1).value for i in range(2, hoja.max_row + 1)]

    # Cierra el libro de trabajo
    libro.close()

    return lectura